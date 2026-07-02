#!/usr/bin/env python3
"""Build a public aggregate mortality anchor from NCHS life-table spreadsheets."""

from __future__ import annotations

import argparse
import json
import tempfile
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_OUTPUT = (
    REPO_ROOT
    / "domains"
    / "c1-boundary-rewriting"
    / "longevity-evidence"
    / "data"
    / "manual"
    / "life_path_public_mortality_anchor.json"
)
NCHS_REPORT_URL = "https://www.cdc.gov/nchs/data/nvsr/nvsr72/nvsr72-12.pdf"
SOURCE_TABLES = {
    "male": {
        "tableId": "nchs-2021-us-male-life-table-table-2",
        "label": "Life table for males: United States, 2021",
        "url": "https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/NVSR/72-12/Table02.xlsx",
    },
    "female": {
        "tableId": "nchs-2021-us-female-life-table-table-3",
        "label": "Life table for females: United States, 2021",
        "url": "https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/NVSR/72-12/Table03.xlsx",
    },
}


def download(url: str, destination: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": "HumanInfraResearch/0.1"})
    with urllib.request.urlopen(request, timeout=45) as response:  # noqa: S310 - public data URL.
        destination.write_bytes(response.read())


def age_from_label(label: Any) -> int | None:
    if not isinstance(label, str):
        return None
    if "and older" in label:
        return int(label.split()[0])
    if "–" in label:
        return int(label.split("–", 1)[0])
    if "-" in label:
        return int(label.split("-", 1)[0])
    return None


def parse_table(path: Path, sex: str, min_age: int, max_age: int) -> dict[str, Any]:
    worksheet = load_workbook(path, data_only=True).active
    title = worksheet.cell(row=1, column=1).value
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        age = age_from_label(row[0])
        if age is None or age < min_age or age > max_age:
            continue
        qx = float(row[1])
        lx = float(row[2])
        dx = float(row[3])
        person_years = float(row[4])
        total_person_years = float(row[5])
        life_expectancy = float(row[6])
        rows.append(
            {
                "age": age,
                "ageInterval": row[0],
                "qx": round(qx, 12),
                "lx": round(lx, 6),
                "dx": round(dx, 6),
                "Lx": round(person_years, 6),
                "Tx": round(total_person_years, 6),
                "ex": round(life_expectancy, 6),
            }
        )
    if len(rows) != max_age - min_age + 1:
        raise ValueError(f"{sex} table produced {len(rows)} rows, expected {max_age - min_age + 1}")
    return {
        "sex": sex,
        "tableId": SOURCE_TABLES[sex]["tableId"],
        "title": title,
        "sourceUrl": SOURCE_TABLES[sex]["url"],
        "columns": {
            "qx": "Probability of dying between ages x and x + 1",
            "lx": "Number surviving to age x",
            "dx": "Number dying between ages x and x + 1",
            "Lx": "Person-years lived between ages x and x + 1",
            "Tx": "Total number of person-years lived above age x",
            "ex": "Expectation of life at age x",
        },
        "rows": rows,
    }


def build_output(min_age: int, max_age: int) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        parsed: list[dict[str, Any]] = []
        for sex, meta in SOURCE_TABLES.items():
            local = temp / f"{sex}.xlsx"
            download(meta["url"], local)
            parsed.append(parse_table(local, sex, min_age, max_age))

    return {
        "schemaVersion": "human-infra.life-path-public-mortality-anchor.v1",
        "status": "public-aggregate-baseline-anchor-not-calibrated-model",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceAuthority": "National Center for Health Statistics, National Vital Statistics System",
        "reportUrl": NCHS_REPORT_URL,
        "sourceTables": [
            {
                "tableId": meta["tableId"],
                "label": meta["label"],
                "url": meta["url"],
            }
            for meta in SOURCE_TABLES.values()
        ],
        "scope": {
            "geography": "United States",
            "periodYear": 2021,
            "ageMin": min_age,
            "ageMax": max_age,
            "sexGroups": ["male", "female"],
            "unit": "aggregate period life table",
        },
        "modelUseBoundary": {
            "allowedUses": [
                "public aggregate mortality baseline anchoring",
                "toy-model baseline plausibility comparison",
                "life-table column contract testing",
            ],
            "blockedUses": [
                "individual prediction",
                "individual death-date output",
                "medical advice",
                "calibrated Human Infra prediction",
                "intervention effect estimation",
                "LEV proof",
            ],
            "noIndividualRows": True,
            "noInterventionEffects": True,
            "noCalibrationClaim": True,
        },
        "tables": parsed,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--min-age", type=int, default=40)
    parser.add_argument("--max-age", type=int, default=100)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.min_age < 0 or args.max_age <= args.min_age:
        raise ValueError("age range must be non-negative and increasing")
    output = build_output(args.min_age, args.max_age)
    output_path = args.out.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(output, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    print(f"wrote {output_path.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
